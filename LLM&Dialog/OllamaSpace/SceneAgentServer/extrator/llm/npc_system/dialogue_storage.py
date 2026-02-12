"""
DialogueStorage - 对话记录存储模块
整合SQLite(.db)和Excel(.xlsx)存储方式

参考原服务器的存储方式:
- SQLChatMessageHistory (memory.db, test_memo.db)
- Excel导出 (history.xlsx, output.xlsx)

功能:
- 对话记录的持久化存储
- 支持SQLite和Excel双重存储
- 会话管理和历史查询
- 与原general_chat.py兼容
"""

import os
import sqlite3
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field, asdict

# 尝试导入openpyxl
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[DialogueStorage] openpyxl未安装，Excel导出功能不可用")


# ==================== 数据结构 ====================

@dataclass
class DialogueMessage:
    """对话消息"""
    message_id: str = ""
    session_id: str = ""
    npc_id: str = ""
    player_id: str = ""
    role: str = "user"  # user/assistant/system
    content: str = ""
    timestamp: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def __post_init__(self):
        if not self.message_id:
            self.message_id = f"msg_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
        if not self.timestamp:
            self.timestamp = datetime.now().isoformat()
    
    def to_dict(self) -> dict:
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: dict) -> 'DialogueMessage':
        return cls(**{k: v for k, v in data.items() if k in cls.__dataclass_fields__})


@dataclass
class DialogueSession:
    """对话会话"""
    session_id: str = ""
    npc_id: str = ""
    player_id: str = ""
    start_time: str = ""
    end_time: str = ""
    message_count: int = 0
    summary: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def __post_init__(self):
        if not self.session_id:
            self.session_id = f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        if not self.start_time:
            self.start_time = datetime.now().isoformat()


# ==================== SQLite存储 ====================

class SQLiteDialogueStore:
    """
    SQLite对话存储
    
    兼容原服务器的SQLChatMessageHistory格式
    """
    
    def __init__(self, db_path: str = "dialogue_history.db"):
        self.db_path = db_path
        self._init_database()
    
    def _init_database(self):
        """初始化数据库"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 会话表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS dialogue_sessions (
                session_id TEXT PRIMARY KEY,
                npc_id TEXT,
                player_id TEXT,
                start_time TEXT,
                end_time TEXT,
                message_count INTEGER DEFAULT 0,
                summary TEXT,
                metadata TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 消息表 - 兼容LangChain的message_store格式
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS dialogue_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                message_id TEXT UNIQUE,
                session_id TEXT,
                npc_id TEXT,
                player_id TEXT,
                role TEXT,
                content TEXT,
                timestamp TEXT,
                metadata TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (session_id) REFERENCES dialogue_sessions(session_id)
            )
        ''')
        
        # 兼容LangChain的message_store表 (用于SQLChatMessageHistory)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS message_store (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT,
                message TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 索引
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_messages_session ON dialogue_messages(session_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_messages_npc ON dialogue_messages(npc_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_messages_player ON dialogue_messages(player_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_messages_timestamp ON dialogue_messages(timestamp)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_langchain_session ON message_store(session_id)')
        
        conn.commit()
        conn.close()
        
        print(f"[SQLiteDialogueStore] 数据库初始化完成: {self.db_path}")
    
    def create_session(self, session: DialogueSession) -> str:
        """创建会话"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR REPLACE INTO dialogue_sessions
            (session_id, npc_id, player_id, start_time, end_time, message_count, summary, metadata)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session.session_id,
            session.npc_id,
            session.player_id,
            session.start_time,
            session.end_time,
            session.message_count,
            session.summary,
            json.dumps(session.metadata, ensure_ascii=False)
        ))
        
        conn.commit()
        conn.close()
        
        return session.session_id
    
    def add_message(self, message: DialogueMessage) -> str:
        """添加消息"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 添加到dialogue_messages表
        cursor.execute('''
            INSERT INTO dialogue_messages
            (message_id, session_id, npc_id, player_id, role, content, timestamp, metadata)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            message.message_id,
            message.session_id,
            message.npc_id,
            message.player_id,
            message.role,
            message.content,
            message.timestamp,
            json.dumps(message.metadata, ensure_ascii=False)
        ))
        
        # 同时添加到message_store表 (兼容LangChain)
        langchain_message = json.dumps({
            "type": message.role,
            "data": {
                "content": message.content,
                "additional_kwargs": message.metadata
            }
        }, ensure_ascii=False)
        
        cursor.execute('''
            INSERT INTO message_store (session_id, message)
            VALUES (?, ?)
        ''', (message.session_id, langchain_message))
        
        # 更新会话消息计数
        cursor.execute('''
            UPDATE dialogue_sessions
            SET message_count = message_count + 1,
                end_time = ?
            WHERE session_id = ?
        ''', (message.timestamp, message.session_id))
        
        conn.commit()
        conn.close()
        
        return message.message_id
    
    def get_session_messages(self, session_id: str, 
                             limit: int = 100,
                             offset: int = 0) -> List[DialogueMessage]:
        """获取会话消息"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT message_id, session_id, npc_id, player_id, role, content, timestamp, metadata
            FROM dialogue_messages
            WHERE session_id = ?
            ORDER BY timestamp ASC
            LIMIT ? OFFSET ?
        ''', (session_id, limit, offset))
        
        rows = cursor.fetchall()
        conn.close()
        
        messages = []
        for row in rows:
            messages.append(DialogueMessage(
                message_id=row[0],
                session_id=row[1],
                npc_id=row[2],
                player_id=row[3],
                role=row[4],
                content=row[5],
                timestamp=row[6],
                metadata=json.loads(row[7]) if row[7] else {}
            ))
        
        return messages
    
    def get_npc_dialogues(self, npc_id: str, 
                          player_id: str = None,
                          limit: int = 50) -> List[DialogueMessage]:
        """获取NPC的对话记录"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        if player_id:
            cursor.execute('''
                SELECT message_id, session_id, npc_id, player_id, role, content, timestamp, metadata
                FROM dialogue_messages
                WHERE npc_id = ? AND player_id = ?
                ORDER BY timestamp DESC
                LIMIT ?
            ''', (npc_id, player_id, limit))
        else:
            cursor.execute('''
                SELECT message_id, session_id, npc_id, player_id, role, content, timestamp, metadata
                FROM dialogue_messages
                WHERE npc_id = ?
                ORDER BY timestamp DESC
                LIMIT ?
            ''', (npc_id, limit))
        
        rows = cursor.fetchall()
        conn.close()
        
        messages = []
        for row in rows:
            messages.append(DialogueMessage(
                message_id=row[0],
                session_id=row[1],
                npc_id=row[2],
                player_id=row[3],
                role=row[4],
                content=row[5],
                timestamp=row[6],
                metadata=json.loads(row[7]) if row[7] else {}
            ))
        
        return messages
    
    def search_dialogues(self, query: str,
                         npc_id: str = None,
                         player_id: str = None,
                         limit: int = 20) -> List[DialogueMessage]:
        """搜索对话记录"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        conditions = ["content LIKE ?"]
        params = [f"%{query}%"]
        
        if npc_id:
            conditions.append("npc_id = ?")
            params.append(npc_id)
        if player_id:
            conditions.append("player_id = ?")
            params.append(player_id)
        
        where_clause = " AND ".join(conditions)
        params.append(limit)
        
        cursor.execute(f'''
            SELECT message_id, session_id, npc_id, player_id, role, content, timestamp, metadata
            FROM dialogue_messages
            WHERE {where_clause}
            ORDER BY timestamp DESC
            LIMIT ?
        ''', params)
        
        rows = cursor.fetchall()
        conn.close()
        
        messages = []
        for row in rows:
            messages.append(DialogueMessage(
                message_id=row[0],
                session_id=row[1],
                npc_id=row[2],
                player_id=row[3],
                role=row[4],
                content=row[5],
                timestamp=row[6],
                metadata=json.loads(row[7]) if row[7] else {}
            ))
        
        return messages
    
    def get_sessions(self, npc_id: str = None,
                     player_id: str = None,
                     limit: int = 20) -> List[DialogueSession]:
        """获取会话列表"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        conditions = []
        params = []
        
        if npc_id:
            conditions.append("npc_id = ?")
            params.append(npc_id)
        if player_id:
            conditions.append("player_id = ?")
            params.append(player_id)
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        params.append(limit)
        
        cursor.execute(f'''
            SELECT session_id, npc_id, player_id, start_time, end_time, message_count, summary, metadata
            FROM dialogue_sessions
            WHERE {where_clause}
            ORDER BY start_time DESC
            LIMIT ?
        ''', params)
        
        rows = cursor.fetchall()
        conn.close()
        
        sessions = []
        for row in rows:
            sessions.append(DialogueSession(
                session_id=row[0],
                npc_id=row[1],
                player_id=row[2],
                start_time=row[3],
                end_time=row[4],
                message_count=row[5],
                summary=row[6],
                metadata=json.loads(row[7]) if row[7] else {}
            ))
        
        return sessions
    
    def get_stats(self) -> Dict[str, Any]:
        """获取统计信息"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) FROM dialogue_sessions')
        session_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM dialogue_messages')
        message_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(DISTINCT npc_id) FROM dialogue_messages')
        npc_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(DISTINCT player_id) FROM dialogue_messages')
        player_count = cursor.fetchone()[0]
        
        conn.close()
        
        return {
            "total_sessions": session_count,
            "total_messages": message_count,
            "unique_npcs": npc_count,
            "unique_players": player_count,
            "db_path": self.db_path
        }
    
    def clear_session(self, session_id: str) -> bool:
        """清除会话"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM dialogue_messages WHERE session_id = ?', (session_id,))
        cursor.execute('DELETE FROM message_store WHERE session_id = ?', (session_id,))
        cursor.execute('DELETE FROM dialogue_sessions WHERE session_id = ?', (session_id,))
        
        conn.commit()
        conn.close()
        
        return True


# ==================== Excel导出 ====================

class ExcelDialogueExporter:
    """
    Excel对话导出器
    
    兼容原服务器的export_history.py格式
    """
    
    def __init__(self, output_dir: str = "./dialogue_exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_session(self, messages: List[DialogueMessage],
                       filename: str = None) -> Optional[str]:
        """导出会话到Excel"""
        if not EXCEL_AVAILABLE:
            print("[ExcelExporter] openpyxl未安装，无法导出Excel")
            return None
        
        if not messages:
            return None
        
        if not filename:
            session_id = messages[0].session_id if messages else "unknown"
            filename = f"dialogue_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = self.output_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "对话记录"
        
        # 表头
        headers = ["消息ID", "会话ID", "NPC", "玩家", "角色", "内容", "时间"]
        ws.append(headers)
        
        # 数据
        for msg in messages:
            ws.append([
                msg.message_id,
                msg.session_id,
                msg.npc_id,
                msg.player_id,
                msg.role,
                msg.content,
                msg.timestamp
            ])
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filepath)
        
        return str(filepath)
    
    def export_all_dialogues(self, sqlite_store: SQLiteDialogueStore,
                             filename: str = "all_dialogues.xlsx") -> Optional[str]:
        """导出所有对话到Excel"""
        if not EXCEL_AVAILABLE:
            return None
        
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        # 会话概览表
        ws_sessions = wb.active
        ws_sessions.title = "会话概览"
        ws_sessions.append(["会话ID", "NPC", "玩家", "开始时间", "结束时间", "消息数", "摘要"])
        
        sessions = sqlite_store.get_sessions(limit=1000)
        for session in sessions:
            ws_sessions.append([
                session.session_id,
                session.npc_id,
                session.player_id,
                session.start_time,
                session.end_time,
                session.message_count,
                session.summary
            ])
        
        # 详细消息表
        ws_messages = wb.create_sheet("详细消息")
        ws_messages.append(["消息ID", "会话ID", "NPC", "玩家", "角色", "内容", "时间"])
        
        for session in sessions:
            messages = sqlite_store.get_session_messages(session.session_id)
            for msg in messages:
                ws_messages.append([
                    msg.message_id,
                    msg.session_id,
                    msg.npc_id,
                    msg.player_id,
                    msg.role,
                    msg.content,
                    msg.timestamp
                ])
        
        wb.save(filepath)
        
        return str(filepath)


# ==================== 统一对话存储管理器 ====================

class DialogueStorage:
    """
    统一对话存储管理器
    
    整合SQLite和Excel存储，提供统一接口
    """
    
    def __init__(self, 
                 db_path: str = "./npc_data/dialogue_history.db",
                 export_dir: str = "./npc_data/dialogue_exports",
                 auto_export: bool = False):
        """
        初始化对话存储
        
        Args:
            db_path: SQLite数据库路径
            export_dir: Excel导出目录
            auto_export: 是否自动导出到Excel
        """
        # 确保目录存在
        Path(db_path).parent.mkdir(parents=True, exist_ok=True)
        
        self.sqlite_store = SQLiteDialogueStore(db_path)
        self.excel_exporter = ExcelDialogueExporter(export_dir)
        self.auto_export = auto_export
        
        # 当前活跃会话
        self.active_sessions: Dict[str, DialogueSession] = {}
    
    def start_session(self, npc_id: str, player_id: str,
                      session_id: str = None) -> DialogueSession:
        """开始新会话"""
        session = DialogueSession(
            session_id=session_id,
            npc_id=npc_id,
            player_id=player_id
        )
        
        self.sqlite_store.create_session(session)
        self.active_sessions[session.session_id] = session
        
        return session
    
    def add_message(self, session_id: str, role: str, content: str,
                    npc_id: str = "", player_id: str = "",
                    metadata: Dict = None) -> DialogueMessage:
        """添加消息"""
        # 获取会话信息
        session = self.active_sessions.get(session_id)
        if session:
            npc_id = npc_id or session.npc_id
            player_id = player_id or session.player_id
        
        message = DialogueMessage(
            session_id=session_id,
            npc_id=npc_id,
            player_id=player_id,
            role=role,
            content=content,
            metadata=metadata or {}
        )
        
        self.sqlite_store.add_message(message)
        
        return message
    
    def add_dialogue_turn(self, session_id: str,
                          player_message: str,
                          npc_reply: str,
                          npc_id: str = "",
                          player_id: str = "",
                          metadata: Dict = None) -> Tuple[DialogueMessage, DialogueMessage]:
        """添加一轮对话 (玩家消息 + NPC回复)"""
        player_msg = self.add_message(
            session_id=session_id,
            role="user",
            content=player_message,
            npc_id=npc_id,
            player_id=player_id,
            metadata=metadata
        )
        
        npc_msg = self.add_message(
            session_id=session_id,
            role="assistant",
            content=npc_reply,
            npc_id=npc_id,
            player_id=player_id,
            metadata=metadata
        )
        
        return player_msg, npc_msg
    
    def end_session(self, session_id: str, summary: str = "") -> Optional[str]:
        """结束会话"""
        if session_id in self.active_sessions:
            del self.active_sessions[session_id]
        
        # 自动导出
        if self.auto_export:
            messages = self.sqlite_store.get_session_messages(session_id)
            return self.excel_exporter.export_session(messages)
        
        return None
    
    def get_history(self, session_id: str = None,
                    npc_id: str = None,
                    player_id: str = None,
                    limit: int = 50) -> List[DialogueMessage]:
        """获取对话历史"""
        if session_id:
            return self.sqlite_store.get_session_messages(session_id, limit)
        elif npc_id:
            return self.sqlite_store.get_npc_dialogues(npc_id, player_id, limit)
        else:
            return []
    
    def search(self, query: str, **kwargs) -> List[DialogueMessage]:
        """搜索对话"""
        return self.sqlite_store.search_dialogues(query, **kwargs)
    
    def export_to_excel(self, session_id: str = None,
                        filename: str = None) -> Optional[str]:
        """导出到Excel"""
        if session_id:
            messages = self.sqlite_store.get_session_messages(session_id)
            return self.excel_exporter.export_session(messages, filename)
        else:
            return self.excel_exporter.export_all_dialogues(
                self.sqlite_store, 
                filename or "all_dialogues.xlsx"
            )
    
    def get_stats(self) -> Dict[str, Any]:
        """获取统计信息"""
        stats = self.sqlite_store.get_stats()
        stats["active_sessions"] = len(self.active_sessions)
        return stats


# ==================== 便捷函数 ====================

def create_dialogue_storage(data_dir: str = "./npc_data") -> DialogueStorage:
    """创建对话存储实例"""
    return DialogueStorage(
        db_path=os.path.join(data_dir, "dialogue_history.db"),
        export_dir=os.path.join(data_dir, "dialogue_exports")
    )
